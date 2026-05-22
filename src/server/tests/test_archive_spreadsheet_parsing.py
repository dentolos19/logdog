from __future__ import annotations

import sys
import zipfile
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from openpyxl import Workbook

from parsers.contracts import ColumnDefinition, ParserPipelineResult, TableDefinition, build_ddl, make_display_name
from parsers.engine import normalize_records
from parsers.orchestrator import _decode_payload, _parse_and_merge
from parsers.preprocessor import LogPreprocessorService
from parsers.registry import ParserRegistry


def test_zip_members_are_parsed_as_separate_parser_runs(monkeypatch):
    """A ZIP with several CSV members should not collapse into the first member."""

    samples_dir = Path(__file__).resolve().parent.parent.parent.parent / "samples"
    zip_path = samples_dir / "lambda.zip"

    file_inputs = _decode_payload("lambda.zip", zip_path.read_bytes())
    assert [file_input.filename for file_input in file_inputs] == [
        "lambda.zip:0.csv",
        "lambda.zip:1.csv",
        "lambda.zip:2.csv",
        "lambda.zip:3.csv",
    ]

    seen_batches: list[list[str]] = []

    class FakeUniversalParser:
        parser_key = "universal_ai"

        def ingest(self, batch, classification):
            seen_batches.append([file_input.filename for file_input in batch])
            assert len(batch) == 1
            file_input = batch[0]
            table_name = f"table_{len(seen_batches)}"
            columns = [ColumnDefinition(name="source", sql_type="TEXT")]
            return ParserPipelineResult(
                table_definitions=[
                    TableDefinition(
                        table_name=table_name,
                        display_name=make_display_name("ai_universal", None, file_input.filename),
                        columns=columns,
                        ddl=build_ddl(table_name, columns),
                    )
                ],
                records={table_name: [{"source": file_input.filename}]},
                parser_key=self.parser_key,
                confidence=0.9,
            )

    def fake_route(parser_key: str):
        assert parser_key == "universal_ai"
        return FakeUniversalParser()

    monkeypatch.setattr(ParserRegistry, "route", fake_route)

    classification = LogPreprocessorService(use_llm=False).classify(file_inputs)
    result = _parse_and_merge(file_inputs=file_inputs, classification=classification)

    assert seen_batches == [[file_input.filename] for file_input in file_inputs]
    assert len(result.table_definitions) == 4
    assert sorted(row["source"] for rows in result.records.values() for row in rows) == sorted(
        file_input.filename for file_input in file_inputs
    )


def test_xlsx_payload_decodes_to_one_csv_input_per_non_empty_sheet():
    workbook = Workbook()
    errors = workbook.active
    errors.title = "Errors"
    errors.append(["timestamp", "level", "message"])
    errors.append(["2026-05-20 00:00:00", "ERROR", "failed"])

    metrics = workbook.create_sheet("Metrics")
    metrics.append(["name", "value"])
    metrics.append(["latency_ms", 12.5])

    workbook.create_sheet("Empty Sheet")

    payload = BytesIO()
    workbook.save(payload)

    file_inputs = _decode_payload("workbook.xlsx", payload.getvalue())

    assert [file_input.filename for file_input in file_inputs] == [
        "workbook.xlsx:Errors.csv",
        "workbook.xlsx:Metrics.csv",
    ]

    errors_records = normalize_records(file_inputs[0].content, file_inputs[0].filename)
    metrics_records = normalize_records(file_inputs[1].content, file_inputs[1].filename)

    assert errors_records == [
        {
            "source": "workbook.xlsx:Errors.csv",
            "record_index": 0,
            "timestamp": "2026-05-20 00:00:00",
            "level": "ERROR",
            "message": "failed",
            "raw": "2026-05-20 00:00:00,ERROR,failed",
        }
    ]
    assert metrics_records[0]["name"] == "latency_ms"
    assert metrics_records[0]["value"] == "12.5"


def test_zip_can_contain_xlsx_members():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Events"
    sheet.append(["id", "message"])
    sheet.append([1, "created"])

    workbook_payload = BytesIO()
    workbook.save(workbook_payload)

    archive_payload = BytesIO()
    with zipfile.ZipFile(archive_payload, "w") as archive:
        archive.writestr("nested/events.xlsx", workbook_payload.getvalue())

    file_inputs = _decode_payload("bundle.zip", archive_payload.getvalue())

    assert [file_input.filename for file_input in file_inputs] == ["bundle.zip:nested/events.xlsx:Events.csv"]
    records = normalize_records(file_inputs[0].content, file_inputs[0].filename)
    assert records[0]["id"] == "1"
    assert records[0]["message"] == "created"
