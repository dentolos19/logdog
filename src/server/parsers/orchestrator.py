from __future__ import annotations

import csv
import gzip
import io
import json
import logging
import os
import tarfile
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, time
from typing import Any

from sqlalchemy.orm import Session

from lib.database import SessionLocal
from lib.megabase import (
    SessionLocal as MegabaseSessionLocal,
    create_table as megabase_create_table,
    init_megabase,
    insert_record as megabase_insert_record,
)
from lib.models import Asset, LogGroup, LogFile, LogProcess, LogTable
from lib.storage import download_file
from parsers.binary import (
    binary_metadata,
    extract_printable_text,
    sha256_bytes,
)
from parsers.contracts import (
    BINARY_PARSER_KEY,
    ClassificationResult,
    ParserPipelineResult,
)
from parsers.normalization import sanitize_db_value
from parsers.preprocessor import FileInput, LogPreprocessorService
from parsers.registry import ParserRegistry

logger = logging.getLogger(__name__)


def _resolve_parse_job_workers() -> int:
    raw_value = os.environ.get("LOG_PARSE_JOB_WORKERS", "4").strip()
    try:
        workers = int(raw_value)
    except ValueError:
        workers = 4
    return max(workers, 1)


PARSE_JOB_EXECUTOR = ThreadPoolExecutor(
    max_workers=_resolve_parse_job_workers(),
    thread_name_prefix="logdog-parse-job",
)

MAX_STORED_PROCESS_RECORDS_PER_TABLE = int(os.environ.get("LOGDOG_MAX_STORED_PROCESS_RECORDS_PER_TABLE", "1000"))
MAX_ARCHIVE_DEPTH = int(os.environ.get("LOGDOG_MAX_ARCHIVE_DEPTH", "2"))
MAX_ARCHIVE_MEMBERS = int(os.environ.get("LOGDOG_MAX_ARCHIVE_MEMBERS", "100"))
MAX_ARCHIVE_MEMBER_SIZE = int(os.environ.get("LOGDOG_MAX_ARCHIVE_MEMBER_SIZE", str(10 * 1024 * 1024)))
MAX_ARCHIVE_TOTAL_SIZE = int(os.environ.get("LOGDOG_MAX_ARCHIVE_TOTAL_SIZE", str(50 * 1024 * 1024)))
MAX_XLSX_ROWS_PER_SHEET = int(os.environ.get("LOGDOG_MAX_XLSX_ROWS_PER_SHEET", "10000"))
MAX_XLSX_CELLS_PER_SHEET = int(os.environ.get("LOGDOG_MAX_XLSX_CELLS_PER_SHEET", "200000"))


def register_pipelines() -> None:
    ParserRegistry.discover(force=True)
    logger.info("Parser pipelines registered: %s", ", ".join(sorted(ParserRegistry.registered_keys())))


def create_process(
    group_id: str,
    file_inputs: list[FileInput] | None = None,
    file_ids: list[str] | None = None,
    file_id: str | None = None,
) -> str:
    db = SessionLocal()
    try:
        group = db.query(LogGroup).filter_by(id=_uuid_or_raw(group_id)).first()
        if group is None:
            raise ValueError(f"Log group '{group_id}' not found.")

        classification_json: str | None = None
        if file_inputs:
            classification = LogPreprocessorService(table_name="logs", profile_name=group.profile_name).classify(
                file_inputs
            )
            classification_json = classification.model_dump_json()

        process = LogProcess(
            group_id=_uuid_or_raw(group_id),
            file_id=_uuid_or_raw(file_id) if file_id else None,
            status="queued",
            classification=classification_json,
        )
        db.add(process)
        db.commit()
        db.refresh(process)

        if file_ids:
            logger.info(
                "Created process %s for group %s with %d file id(s).",
                process.id,
                group_id,
                len(file_ids),
            )

        return str(process.id)
    finally:
        db.close()


def enqueue_process(
    process_id: str,
    group_id: str,
    file_inputs_json: str | None = None,
    file_ids_json: str | None = None,
) -> None:
    PARSE_JOB_EXECUTOR.submit(
        run_parse_job,
        process_id,
        group_id,
        file_inputs_json,
        file_ids_json,
    )


def mark_process_failed(process_id: str, group_id: str, message: str) -> None:
    db = SessionLocal()
    try:
        process = db.query(LogProcess).filter_by(id=_uuid_or_raw(process_id), group_id=_uuid_or_raw(group_id)).first()
        _fail(db=db, process=process, message=message)
    finally:
        db.close()


def orchestrate_files(
    group_id: str,
    file_inputs: list[FileInput],
    persist: bool = True,
    use_llm: bool = True,
) -> ParserPipelineResult:
    db = SessionLocal()
    megabase_db = MegabaseSessionLocal()
    try:
        register_pipelines()
        init_megabase(megabase_db)

        group = db.query(LogGroup).filter_by(id=_uuid_or_raw(group_id)).first()
        profile_name = group.profile_name if group is not None else "default"
        preprocessor = LogPreprocessorService(table_name="logs", use_llm=use_llm, profile_name=profile_name)
        classification = preprocessor.classify(file_inputs)

        pipeline_result = _parse_and_merge(file_inputs=file_inputs, classification=classification)

        if persist:
            _persist_artifacts(db=db, megabase_db=megabase_db, group_id=group_id, result=pipeline_result)

        return pipeline_result
    finally:
        megabase_db.close()
        db.close()


def run_parse_job(
    process_id: str,
    group_id: str,
    file_inputs_json: str | None = None,
    file_ids_json: str | None = None,
) -> None:
    db = SessionLocal()
    megabase_db = MegabaseSessionLocal()

    try:
        init_megabase(megabase_db)
        register_pipelines()

        process = db.query(LogProcess).filter_by(id=_uuid_or_raw(process_id), group_id=_uuid_or_raw(group_id)).first()
        if process is None:
            logger.error("run_parse_job: process %s not found for group %s", process_id, group_id)
            return

        process.status = "processing"
        process.error = None
        db.commit()

        group = db.query(LogGroup).filter_by(id=_uuid_or_raw(group_id)).first()
        profile_name = group.profile_name if group is not None else "default"

        file_inputs = _resolve_file_inputs(
            db=db,
            group_id=group_id,
            file_inputs_json=file_inputs_json,
            file_ids_json=file_ids_json,
        )
        if not file_inputs:
            _fail(db, process, "No file inputs available to parse.")
            return

        preprocessor = LogPreprocessorService(table_name="logs", use_llm=True, profile_name=profile_name)
        classification = preprocessor.classify(file_inputs)
        process.classification = classification.model_dump_json()
        db.commit()

        pipeline_result = _parse_and_merge(file_inputs=file_inputs, classification=classification)
        if not pipeline_result.table_definitions:
            _fail(db, process, "; ".join(pipeline_result.warnings) or "No tables were produced.")
            return

        _persist_artifacts(db=db, megabase_db=megabase_db, group_id=group_id, result=pipeline_result)

        safe_result = _make_json_safe_pipeline_result(pipeline_result)
        process.result = json.dumps(safe_result, ensure_ascii=True, default=str)
        process.status = "completed"
        process.error = None
        db.commit()
        logger.info("run_parse_job: process=%s completed", process_id)
    except Exception as error:  # noqa: BLE001
        logger.exception("run_parse_job: unhandled error for process %s", process_id)
        process = db.query(LogProcess).filter_by(id=_uuid_or_raw(process_id)).first()
        _fail(db, process, str(error))
    finally:
        megabase_db.close()
        db.close()


# ── File input resolution ────────────────────────────────────────────────


def _resolve_file_inputs(
    db: Session,
    group_id: str,
    file_inputs_json: str | None,
    file_ids_json: str | None,
) -> list[FileInput]:
    if file_inputs_json:
        raw_inputs: list[dict[str, Any]] = json.loads(file_inputs_json)
        return [FileInput(**item) for item in raw_inputs]

    file_id_filter: set[str] | None = None
    if file_ids_json:
        parsed_ids = json.loads(file_ids_json)
        file_id_filter = {str(value) for value in parsed_ids}

    file_rows = db.query(LogFile).filter_by(group_id=_uuid_or_raw(group_id)).all()
    if file_id_filter:
        file_rows = [row for row in file_rows if str(row.id) in file_id_filter]

    file_inputs: list[FileInput] = []
    for file_row in file_rows:
        asset = db.query(Asset).filter_by(id=file_row.asset_id).first()
        if asset is None:
            logger.warning("Skipping log_file %s because linked asset is missing.", file_row.id)
            continue

        raw_bytes = download_file(file_row.asset_id, db=db)
        if raw_bytes is None:
            logger.warning("Skipping file %s because storage payload could not be downloaded.", asset.name)
            continue

        decoded_members = _decode_payload(asset.name, raw_bytes)
        for member_input in decoded_members:
            member_input.file_id = str(file_row.id)
            file_inputs.append(member_input)

    return file_inputs


# ── Archive / file decode helpers ────────────────────────────────────────


def _decode_bytes(raw_bytes: bytes, filename: str) -> str:
    if _is_hex_dump(raw_bytes):
        return _decode_hex_dump(raw_bytes)
    try:
        return raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return raw_bytes.decode("latin-1")
        except UnicodeDecodeError:
            return raw_bytes.decode("utf-8", errors="ignore")


def _decode_payload_to_file_input(filename: str, raw_bytes: bytes) -> FileInput:
    """Decode *raw_bytes* into a ``FileInput``, preserving binary data.

    For text-like content, produces a normal ``FileInput`` with decoded text.
    For binary content, preserves ``raw_bytes``, extracts printable strings
    as ``content``, and sets binary metadata.
    """
    meta = binary_metadata(raw_bytes, filename)
    is_binary = meta["is_binary"]

    if is_binary:
        extracted_text = extract_printable_text(raw_bytes)
        return FileInput(
            filename=filename,
            content=extracted_text,
            raw_bytes=raw_bytes,
            is_binary=True,
            byte_length=meta["byte_length"],
            sha256=meta["sha256"],
        )

    # Text content — decode normally
    decoded = _decode_bytes(raw_bytes, filename)
    return FileInput(
        filename=filename,
        content=decoded,
        raw_bytes=None,
        is_binary=False,
        byte_length=len(raw_bytes),
        sha256=sha256_bytes(raw_bytes),
    )


def _is_hex_dump(raw_bytes: bytes) -> bool:
    try:
        text = raw_bytes.decode("ascii", errors="ignore")
        lines = text.splitlines()[:10]
        hex_lines = 0
        for line in lines:
            parts = line.split()
            if len(parts) >= 2 and all(
                len(p) == 2 and all(c in "0123456789abcdefABCDEF" for c in p) for p in parts[:8]
            ):
                hex_lines += 1
        return hex_lines >= 3
    except Exception:
        return False


def _decode_hex_dump(raw_bytes: bytes) -> str:
    try:
        text = raw_bytes.decode("ascii", errors="ignore")
        result = []
        for line in text.splitlines():
            parts = line.split()
            hex_part = ""
            for p in parts:
                if len(p) == 2 and all(c in "0123456789abcdefABCDEF" for c in p):
                    hex_part += p
            if hex_part:
                try:
                    decoded = bytes.fromhex(hex_part).decode("utf-8", errors="ignore")
                    if decoded.strip():
                        result.append(decoded)
                except ValueError:
                    pass
        return "\n".join(result) if result else text
    except Exception:
        return raw_bytes.decode("utf-8", errors="ignore")


def _decode_payload(filename: str, raw_bytes: bytes, *, depth: int = 0) -> list[FileInput]:
    """Decode *raw_bytes* into one or more ``FileInput`` objects.

    Spreadsheet workbooks are converted into per-sheet CSV-like text inputs.
    Archives (ZIP, GZIP, tar) are expanded into per-member ``FileInput``
    objects. Plain files produce a single ``FileInput``. Each member is
    independently classified as text or binary.
    """
    if depth > MAX_ARCHIVE_DEPTH:
        logger.warning("Archive depth limit reached while decoding %s", filename)
        return [_decode_payload_to_file_input(filename, raw_bytes)]

    workbook_inputs = _decode_xlsx_payload(filename, raw_bytes)
    if workbook_inputs:
        return workbook_inputs

    archive_members = _extract_archive_members(filename, raw_bytes)
    if archive_members:
        expanded: list[FileInput] = []
        for member_name, member_bytes in archive_members:
            synthetic_name = f"{filename}:{member_name}"
            expanded.extend(_decode_payload(synthetic_name, member_bytes, depth=depth + 1))
        return expanded
    return [_decode_payload_to_file_input(filename, raw_bytes)]


def _decode_xlsx_payload(filename: str, raw_bytes: bytes) -> list[FileInput]:
    """Convert an Excel workbook into one text ``FileInput`` per sheet.

    XLSX files are ZIP containers internally, so this must run before generic
    archive expansion.  Each non-empty worksheet is serialized to CSV text and
    then parsed by the normal CSV-aware parser path.
    """

    if not _looks_like_xlsx(filename, raw_bytes):
        return []

    try:
        from openpyxl import load_workbook
    except Exception as error:  # noqa: BLE001
        logger.warning("openpyxl is unavailable; cannot decode workbook %s: %s", filename, error)
        return []

    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001
        logger.warning("Could not decode workbook %s: %s", filename, error)
        return []

    try:
        inputs: list[FileInput] = []
        for worksheet in workbook.worksheets:
            rows = []
            cell_count = 0
            for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if index > MAX_XLSX_ROWS_PER_SHEET:
                    logger.warning(
                        "Workbook sheet %s was truncated at %d rows", worksheet.title, MAX_XLSX_ROWS_PER_SHEET
                    )
                    break
                cell_count += len(row)
                if cell_count > MAX_XLSX_CELLS_PER_SHEET:
                    logger.warning(
                        "Workbook sheet %s was truncated at %d cells", worksheet.title, MAX_XLSX_CELLS_PER_SHEET
                    )
                    break
                rows.append(row)
            trimmed_rows = _trim_empty_xlsx_rows(rows)
            if not trimmed_rows:
                continue

            csv_content = _xlsx_rows_to_csv(trimmed_rows)
            sheet_filename = f"{filename}:{worksheet.title}.csv"
            encoded = csv_content.encode("utf-8")
            inputs.append(
                FileInput(
                    filename=sheet_filename,
                    content=csv_content,
                    raw_bytes=None,
                    is_binary=False,
                    encoding="utf-8",
                    mime_type="text/csv",
                    byte_length=len(encoded),
                    sha256=sha256_bytes(encoded),
                )
            )
        return inputs
    finally:
        workbook.close()


def _looks_like_xlsx(filename: str, raw_bytes: bytes) -> bool:
    lowered = filename.lower()
    if lowered.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return True
    if not raw_bytes.startswith(b"PK\x03\x04"):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
            names = set(zf.namelist())
    except zipfile.BadZipFile:
        return False
    return "[Content_Types].xml" in names and any(name.startswith("xl/") for name in names)


def _trim_empty_xlsx_rows(rows: list[tuple[Any, ...]]) -> list[list[Any]]:
    non_empty_rows = [list(row) for row in rows if any(_xlsx_cell_has_value(value) for value in row)]
    if not non_empty_rows:
        return []

    max_non_empty_col = 0
    for row in non_empty_rows:
        for index, value in enumerate(row):
            if _xlsx_cell_has_value(value):
                max_non_empty_col = max(max_non_empty_col, index + 1)

    return [row[:max_non_empty_col] for row in non_empty_rows]


def _xlsx_cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _xlsx_rows_to_csv(rows: list[list[Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow([_serialize_xlsx_cell(value) for value in row])
    return output.getvalue()


def _serialize_xlsx_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def _extract_archive_members(filename: str, raw_bytes: bytes) -> list[tuple[str, bytes]]:
    total_size = 0

    def can_accept_member(size: int) -> bool:
        return size <= MAX_ARCHIVE_MEMBER_SIZE and total_size + size <= MAX_ARCHIVE_TOTAL_SIZE

    if raw_bytes.startswith(b"PK\x03\x04"):
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
            members: list[tuple[str, bytes]] = []
            for info in zf.infolist():
                if len(members) >= MAX_ARCHIVE_MEMBERS:
                    break
                if info.is_dir():
                    continue
                if not can_accept_member(info.file_size):
                    continue
                member_bytes = zf.read(info.filename)
                if not can_accept_member(len(member_bytes)):
                    continue
                total_size += len(member_bytes)
                members.append((info.filename, member_bytes))
            return members

    if raw_bytes.startswith(b"\x1f\x8b\x08"):
        output = io.BytesIO()
        with gzip.GzipFile(fileobj=io.BytesIO(raw_bytes)) as gz:
            while output.tell() <= MAX_ARCHIVE_MEMBER_SIZE:
                chunk = gz.read(1024 * 1024)
                if not chunk:
                    break
                output.write(chunk)
        decompressed = output.getvalue()[:MAX_ARCHIVE_MEMBER_SIZE]
        base_name = filename[:-3] if filename.lower().endswith(".gz") else f"{filename}.decompressed"
        return [(base_name, decompressed)]

    if len(raw_bytes) > 262 and raw_bytes[257:262] == b"ustar":
        members = []
        with tarfile.open(fileobj=io.BytesIO(raw_bytes), mode="r:*") as tf:
            for member in tf.getmembers():
                if len(members) >= MAX_ARCHIVE_MEMBERS:
                    break
                if not member.isfile():
                    continue
                if not can_accept_member(member.size):
                    continue
                extracted = tf.extractfile(member)
                if extracted is None:
                    continue
                member_bytes = extracted.read(MAX_ARCHIVE_MEMBER_SIZE + 1)
                if not can_accept_member(len(member_bytes)):
                    continue
                total_size += len(member_bytes)
                members.append((member.name, member_bytes))
        return members

    return []


# ── Parse and merge with universal AI parser + raw fallback ──────────────


def _make_json_safe_pipeline_result(result: ParserPipelineResult) -> dict[str, Any]:
    """Convert a ``ParserPipelineResult`` to a JSON-safe dict.

    Replaces any raw ``bytes`` in records with metadata summaries
    (byte length, SHA-256, truncation flag) to avoid serialization
    failures when storing ``process.result``.
    """
    data = result.model_dump()
    for table_name, records in data.get("records", {}).items():
        original_count = len(records)
        if original_count > MAX_STORED_PROCESS_RECORDS_PER_TABLE:
            data.setdefault("diagnostics", {}).setdefault("truncated_records", {})[table_name] = {
                "stored": MAX_STORED_PROCESS_RECORDS_PER_TABLE,
                "total": original_count,
            }
            del records[MAX_STORED_PROCESS_RECORDS_PER_TABLE:]
        for record in records:
            for key, value in list(record.items()):
                if isinstance(value, bytes):
                    record[key] = {
                        "_binary": True,
                        "byte_length": len(value),
                        "sha256": sha256_bytes(value),
                        "truncated": False,
                    }
    return data


def _parse_and_merge(
    file_inputs: list[FileInput],
    classification: ClassificationResult,
) -> ParserPipelineResult:
    """Parse all files using the appropriate parser for each file type.

    - Binary files (``is_binary=True``) are routed to ``binary_file`` parser.
    - Text files are parsed with ``universal_ai``, falling back to ``raw_ingest``.

    Confidence aggregation uses conservative row-weighted scoring:
      - Each parser's confidence is weighted by its row contribution.
      - Fallback usage incurs a penalty.
      - If raw_ingest contributes >=50% of rows the result is capped at 0.55.
      - If raw_ingest contributes >=80% of rows the result is capped at 0.45.
    """
    merged_table_definitions: list[Any] = []
    merged_records: dict[str, list[dict[str, Any]]] = {}
    merged_warnings: list[str] = []
    used_parser_keys: list[str] = []
    # Track per-parser confidence and row counts for weighted aggregation
    parser_confidences: list[float] = []
    parser_row_counts: list[int] = []
    merged_diagnostics: dict[str, Any] = {
        "parsers": {},
        "fallbacks": [],
        "parser_runs": [],
        "table_row_counts": {},
    }

    def merge_parser_result(
        result: ParserPipelineResult,
        *,
        source_filename: str,
        fallback: bool = False,
    ) -> int:
        """Merge a single parser run into the aggregate result.

        ZIP and tar archives expand to multiple ``FileInput`` objects.  Parser
        runs are merged by table name so each member can produce its own table
        without overwriting tables from previous members.
        """

        merged_table_definitions.extend(result.table_definitions)
        for table_name, rows in result.records.items():
            merged_records.setdefault(table_name, []).extend(rows)
        merged_warnings.extend(result.warnings)
        used_parser_keys.append(result.parser_key)

        row_count = sum(len(rows) for rows in result.records.values())
        parser_confidences.append(result.confidence)
        parser_row_counts.append(row_count)

        parser_run = {
            "parser_key": result.parser_key,
            "source": source_filename,
            "tables": list(result.records.keys()),
            "row_count": row_count,
            "fallback": fallback,
            "diagnostics": result.diagnostics or {},
        }
        merged_diagnostics["parser_runs"].append(parser_run)
        parser_bucket = merged_diagnostics["parsers"].setdefault(result.parser_key, {"runs": []})
        if isinstance(parser_bucket, dict) and isinstance(parser_bucket.get("runs"), list):
            parser_bucket["runs"].append(parser_run)

        return row_count

    # ── Split binary vs. text files ──────────────────────────────────
    binary_inputs: list[FileInput] = []
    text_inputs: list[FileInput] = []
    for fi in file_inputs:
        if fi.is_binary:
            binary_inputs.append(fi)
        else:
            text_inputs.append(fi)

    fallback_used = False  # tracks whether raw_ingest fallback was used
    raw_ingest_row_count = 0

    # ── Parse binary files with BinaryFileParser ─────────────────────
    if binary_inputs:
        try:
            binary_pipeline = ParserRegistry.route(BINARY_PARSER_KEY)
            binary_result = binary_pipeline.ingest(binary_inputs, classification)
            merge_parser_result(
                binary_result,
                source_filename=", ".join(file_input.filename for file_input in binary_inputs),
            )
        except Exception as error:  # noqa: BLE001
            logger.exception("Binary file parser failed")
            merged_warnings.append(f"Binary file parser failed: {error}")
            merged_diagnostics["fallbacks"].append(
                {
                    "from_parser": BINARY_PARSER_KEY,
                    "to_parser": "(none)",
                    "reason": str(error),
                }
            )

    # ── Parse text files with universal AI parser ────────────────────
    # Process each decoded member independently.  Passing every ZIP member to
    # one AI parser run lets the schema/extraction sample be dominated by the
    # first member and can collapse all rows into a single table named after
    # 0.csv.  Per-file runs preserve every archive member and allow mixed
    # schemas to naturally produce multiple tables.
    if text_inputs:
        ai_pipeline = ParserRegistry.route("universal_ai")
        fallback_pipeline = None

        for text_input in text_inputs:
            ai_result = None
            ai_error: Exception | None = None

            try:
                ai_result = ai_pipeline.ingest([text_input], classification)
            except Exception as error:  # noqa: BLE001
                ai_error = error
                logger.exception("Universal AI parser failed for %s", text_input.filename)
                merged_warnings.append(f"Universal AI parser failed for {text_input.filename}: {error}")

            ai_has_rows = ai_result is not None and any(ai_result.records.values())
            if ai_has_rows:
                merge_parser_result(ai_result, source_filename=text_input.filename)
                continue

            if ai_result is not None:
                merged_warnings.extend(ai_result.warnings)

            fallback_used = True
            fallback_reason = "AI parser produced no rows" if ai_result is not None else "AI parser failed"
            merged_diagnostics["fallbacks"].append(
                {
                    "source": text_input.filename,
                    "from_parser": "universal_ai" if ai_result is not None else "(none)",
                    "to_parser": "raw_ingest",
                    "reason": str(ai_error) if ai_error is not None else fallback_reason,
                }
            )

            try:
                if fallback_pipeline is None:
                    fallback_pipeline = ParserRegistry.route("raw_ingest")
                fallback_result = fallback_pipeline.ingest([text_input], classification)
                raw_ingest_row_count += merge_parser_result(
                    fallback_result,
                    source_filename=text_input.filename,
                    fallback=True,
                )
            except Exception as fallback_error:  # noqa: BLE001
                logger.exception("Raw ingest fallback also failed for %s", text_input.filename)
                merged_warnings.append(f"Raw ingest fallback also failed for {text_input.filename}: {fallback_error}")

    # ── Conservative confidence aggregation ──────────────────────────
    total_rows = sum(parser_row_counts)

    if total_rows == 0:
        final_confidence = 0.0
    elif not parser_confidences:
        final_confidence = 0.0
    else:
        # Row-weighted average of parser confidences
        weighted_confidence = (
            sum(conf * count for conf, count in zip(parser_confidences, parser_row_counts)) / total_rows
        )

        # Determine raw_ingest row ratio (if fallback was used)
        raw_ingest_row_ratio = 0.0
        fallback_penalty = 0.0
        applied_cap = None

        if fallback_used:
            raw_ingest_row_ratio = raw_ingest_row_count / total_rows if total_rows > 0 else 0.0
            # Fallback penalty scales with how much of the data came from raw_ingest.
            fallback_penalty = 0.15 + 0.25 * raw_ingest_row_ratio

        final_confidence = weighted_confidence - fallback_penalty

        # Apply caps for fallback-heavy results
        if raw_ingest_row_ratio >= 0.80:
            final_confidence = min(final_confidence, 0.45)
            applied_cap = 0.45
        elif raw_ingest_row_ratio >= 0.50:
            final_confidence = min(final_confidence, 0.55)
            applied_cap = 0.55

        final_confidence = max(0.0, min(1.0, final_confidence))

        merged_diagnostics["confidence_aggregation"] = {
            "formula_version": "orchestrator-v1",
            "weighted_confidence": round(weighted_confidence, 4),
            "fallback_used": fallback_used,
            "raw_ingest_row_ratio": round(raw_ingest_row_ratio, 4),
            "fallback_penalty": round(fallback_penalty, 4),
            "applied_cap": applied_cap,
            "final_confidence": round(final_confidence, 4),
        }

    final_parser_key = "mixed"
    if len(set(used_parser_keys)) == 1 and used_parser_keys:
        final_parser_key = used_parser_keys[0]

    merged_diagnostics["parser_used"] = final_parser_key
    merged_diagnostics["row_counts"] = {table_name: len(rows) for table_name, rows in merged_records.items()}

    return ParserPipelineResult(
        table_definitions=merged_table_definitions,
        records=merged_records,
        parser_key=final_parser_key,
        warnings=merged_warnings,
        confidence=round(final_confidence, 2),
        diagnostics=merged_diagnostics,
    )


# ── Megabase persistence ─────────────────────────────────────────────────


def _persist_artifacts(
    db: Session,
    megabase_db: Session,
    group_id: str,
    result: ParserPipelineResult,
) -> None:
    for table_definition in result.table_definitions:
        rows = result.records.get(table_definition.table_name, [])
        if not rows:
            logger.debug("Skipping empty table=%s (0 rows)", table_definition.table_name)
            continue
        _ensure_megabase_table(megabase_db, table_definition)
        inserted = _insert_rows(megabase_db, table_definition, rows)
        logger.debug("Persisted table=%s rows=%d", table_definition.table_name, inserted)
        _sync_log_table(db=db, group_id=group_id, table_definition=table_definition)


def _ensure_megabase_table(megabase_db: Session, table_definition: Any) -> None:
    schema = {
        "columns": [
            {
                "name": column.name,
                "type": _sql_to_megabase_type(column.sql_type),
                "nullable": column.nullable,
                "primary_key": column.primary_key,
                "description": column.description or "",
            }
            for column in table_definition.columns
        ]
    }

    try:
        megabase_create_table(megabase_db, table_definition.table_name, schema)
    except ValueError as error:
        if "already exists" not in str(error).lower():
            raise


def _insert_rows(
    megabase_db: Session,
    table_definition: Any,
    rows: list[dict[str, Any]],
) -> int:
    allowed_columns = {column.name for column in table_definition.columns}
    ts_columns = {
        column.name
        for column in table_definition.columns
        if column.sql_type.upper() in {"TIMESTAMP", "DATETIME", "TIMESTAMPTZ"}
    }
    inserted = 0
    for row in rows:
        payload = {}
        for key, value in row.items():
            if key not in allowed_columns:
                continue
            if key in ts_columns:
                payload[key] = _normalize_timestamp_value(value)
            else:
                payload[key] = _normalize_value(value)
        megabase_insert_record(megabase_db, table_definition.table_name, payload)
        inserted += 1
    return inserted


def _normalize_value(value: Any) -> Any:
    sanitized = sanitize_db_value(value)
    if isinstance(sanitized, (dict, list)):
        return json.dumps(sanitized, ensure_ascii=True)
    return sanitized


def _normalize_timestamp_value(value: Any) -> Any:
    """Convert a timestamp value to a timezone-aware datetime for Postgres."""
    from parsers.normalization import parse_timestamp

    if value is None or value == "":
        return None
    parsed = parse_timestamp(value)
    if parsed is not None:
        return parsed
    # If parsing fails, return None rather than a broken string
    return None


def _sql_to_megabase_type(sql_type: str) -> str:
    normalized = sql_type.upper()
    if normalized in {"INTEGER", "INT", "SMALLINT"}:
        return "integer"
    if normalized in {"BIGINT"}:
        return "bigint"
    if normalized in {"REAL", "FLOAT", "DOUBLE", "NUMERIC", "DECIMAL"}:
        return "float"
    if normalized in {"BOOLEAN", "BOOL"}:
        return "boolean"
    if normalized in {"JSON", "JSONB"}:
        return "json"
    if normalized in {"BYTEA", "BLOB", "BINARY"}:
        return "bytea"
    if normalized in {"DATETIME", "TIMESTAMP", "TIMESTAMPTZ"} or "TIMESTAMP" in normalized:
        return "datetime"
    if normalized in {"VARCHAR", "CHAR", "STRING"}:
        return "string"
    return "text"


def _sync_log_table(db: Session, group_id: str, table_definition: Any) -> None:
    schema_json = json.dumps(
        [
            {
                "name": column.name,
                "type": column.sql_type,
                "nullable": column.nullable,
                "primary_key": column.primary_key,
                "description": column.description,
            }
            for column in table_definition.columns
        ],
        ensure_ascii=True,
    )

    table_uuid = uuid.UUID(table_definition.table_name)
    existing = db.query(LogTable).filter_by(table=table_definition.table_name).first()
    if existing:
        existing.name = table_definition.display_name
        existing.table = table_definition.table_name
        existing.schema = schema_json
    else:
        db.add(
            LogTable(
                id=table_uuid,
                group_id=_uuid_or_raw(group_id),
                name=table_definition.display_name,
                table=table_definition.table_name,
                schema=schema_json,
            )
        )

    db.commit()


def _uuid_or_raw(value: str) -> Any:
    try:
        return uuid.UUID(str(value))
    except ValueError:
        return value


def _fail(db: Session, process: LogProcess | None, message: str) -> None:
    if process is None:
        return
    try:
        process.status = "failed"
        process.error = message
        db.commit()
    except Exception:  # noqa: BLE001
        logger.exception("Could not persist failure for process %s", getattr(process, "id", "?"))


def get_pipeline_stats() -> dict[str, Any]:
    return {
        "registered_parsers": sorted(ParserRegistry.registered_keys()),
    }
